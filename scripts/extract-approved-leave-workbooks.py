#!/usr/bin/env python3

import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook

LEAVE_FILL_RGB = "FFFF0000"
EMPLOYEE_NAME_COLUMN = 2
MONTH_ROW = 4
DAY_NUMBER_ROW = 5
EMPLOYEE_START_ROW = 8


def normalized_fill_rgb(cell):
    fill = getattr(cell, "fill", None)
    if fill is None or fill.fill_type != "solid":
        return None

    color = getattr(fill, "fgColor", None)
    rgb = getattr(color, "rgb", None)

    if rgb is None:
        color = getattr(fill, "start_color", None)
        rgb = getattr(color, "rgb", None)

    return str(rgb).upper() if rgb else None


def build_date_by_column(worksheet):
    current_year = None
    current_month = None
    date_by_column = {}

    for column in range(1, worksheet.max_column + 1):
        anchor_value = worksheet.cell(row=MONTH_ROW, column=column).value

        if isinstance(anchor_value, datetime):
            current_year = anchor_value.year
            current_month = anchor_value.month
        elif isinstance(anchor_value, date):
            current_year = anchor_value.year
            current_month = anchor_value.month

        if current_year is None or current_month is None:
            continue

        day_value = worksheet.cell(row=DAY_NUMBER_ROW, column=column).value

        if isinstance(day_value, bool) or not isinstance(day_value, (int, float)):
            continue

        if int(day_value) != day_value:
            continue

        try:
            current_date = date(current_year, current_month, int(day_value))
        except ValueError:
            continue

        date_by_column[column] = current_date.isoformat()

    return date_by_column


def find_legend_row(worksheet):
    for row in range(1, worksheet.max_row + 1):
        for column in (1, 2, 3, 4):
            value = worksheet.cell(row=row, column=column).value
            if isinstance(value, str) and value.strip().lower() == "legend":
                return row

    return worksheet.max_row + 1


def extract_workbook(path):
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook.active
    legend_row = find_legend_row(worksheet)
    date_by_column = build_date_by_column(worksheet)

    employees = []

    for row in range(EMPLOYEE_START_ROW, legend_row):
        raw_name = worksheet.cell(row=row, column=EMPLOYEE_NAME_COLUMN).value

        if not isinstance(raw_name, str) or not raw_name.strip():
            continue

        dates = []

        for column, iso_date in date_by_column.items():
            cell = worksheet.cell(row=row, column=column)
            if normalized_fill_rgb(cell) == LEAVE_FILL_RGB:
                dates.append(iso_date)

        if dates:
            employees.append(
                {
                    "name": raw_name.strip(),
                    "row": row,
                    "dates": sorted(set(dates)),
                }
            )

    return {
        "sourcePath": path,
        "sheetName": worksheet.title,
        "employees": employees,
    }


def main(argv):
    paths = argv[1:]

    if not paths:
        print("Usage: extract-approved-leave-workbooks.py <xlsx path> [<xlsx path> ...]", file=sys.stderr)
        return 1

    payload = [extract_workbook(path) for path in paths]
    json.dump(payload, sys.stdout)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
